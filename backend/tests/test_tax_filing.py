"""Tax filing pack and export tests."""

from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app import models as m
from app import tax as tax_svc
from app.report_export import EXPORTABLE, export_report, flatten_report


@pytest.mark.asyncio
async def test_tax_filing_pack_boxes_and_schedules(db_session, seeded):
    tenant_id = seeded["t1"].id
    party = m.Party(tenant_id=tenant_id, name="Buyer", kind="customer", credit_limit=0)
    supplier = m.Party(tenant_id=tenant_id, name="Vendor", kind="supplier", credit_limit=0)
    db_session.add_all([party, supplier])
    await db_session.flush()

    inv = m.SalesInvoice(
        tenant_id=tenant_id,
        invoice_number="INV-TAX-1",
        customer_id=party.id,
        status="posted",
        subtotal=200,
        tax_amount=30,
        total_amount=230,
        paid_amount=0,
        posted_at=datetime(2026, 3, 15),
        created_by=seeded["admin1"].id,
    )
    pi = m.PurchaseInvoice(
        tenant_id=tenant_id,
        invoice_number="PI-TAX-1",
        supplier_id=supplier.id,
        status="unpaid",
        subtotal=100,
        tax_amount=15,
        total_amount=115,
        paid_amount=0,
        invoice_date=datetime(2026, 3, 10),
        supplier_invoice_number="SUP-99",
        created_by=seeded["admin1"].id,
    )
    db_session.add_all([inv, pi])
    await db_session.commit()

    pack = await tax_svc.tax_filing_pack(
        db_session,
        tenant_id,
        from_date=datetime(2026, 3, 1),
        to_date=datetime(2026, 3, 31, 23, 59, 59),
    )
    assert pack["output_tax"] == 30.0
    assert pack["input_tax"] == 15.0
    assert pack["net_tax_payable"] == 15.0
    assert pack["input_tax_source"] == "purchase_invoices"
    boxes = {b["code"]: b["amount"] for b in pack["filing_boxes"]["boxes"]}
    assert boxes["taxable_outputs_net"] == 200.0
    assert boxes["output_tax"] == 30.0
    assert boxes["input_tax"] == 15.0
    assert boxes["net_tax_payable"] == 15.0
    assert len(pack["schedules"]["output"]) == 1
    assert pack["schedules"]["output"][0]["document_number"] == "INV-TAX-1"
    assert len(pack["schedules"]["input"]) == 1
    assert pack["schedules"]["input"][0]["supplier_invoice_number"] == "SUP-99"


@pytest.mark.asyncio
async def test_tax_filing_splits_standard_zero_exempt_supplies(db_session, seeded):
    tenant_id = seeded["t1"].id
    party = m.Party(tenant_id=tenant_id, name="Split Buyer", kind="customer", credit_limit=0)
    db_session.add(party)
    await db_session.flush()

    product_std = m.Product(
        tenant_id=tenant_id,
        name="Standard Good",
        sku="STD-1",
        selling_price=100,
        cost_price=40,
        tax_exempt=False,
    )
    product_zero = m.Product(
        tenant_id=tenant_id,
        name="Zero Rated Export",
        sku="ZERO-1",
        selling_price=80,
        cost_price=30,
        tax_exempt=False,
    )
    product_ex = m.Product(
        tenant_id=tenant_id,
        name="Exempt Essential",
        sku="EX-1",
        selling_price=50,
        cost_price=20,
        tax_exempt=True,
    )
    db_session.add_all([product_std, product_zero, product_ex])
    await db_session.flush()

    inv = m.SalesInvoice(
        tenant_id=tenant_id,
        invoice_number="INV-SPLIT-1",
        customer_id=party.id,
        status="posted",
        subtotal=230,
        tax_amount=15,
        total_amount=245,
        paid_amount=0,
        posted_at=datetime(2026, 4, 10),
        created_by=seeded["admin1"].id,
    )
    db_session.add(inv)
    await db_session.flush()
    db_session.add_all(
        [
            m.SalesInvoiceItem(
                tenant_id=tenant_id,
                sales_invoice_id=inv.id,
                product_id=product_std.id,
                quantity=1,
                unit_price=100,
                tax_rate=15,
                line_total=115,
                supply_category="standard",
            ),
            m.SalesInvoiceItem(
                tenant_id=tenant_id,
                sales_invoice_id=inv.id,
                product_id=product_zero.id,
                quantity=1,
                unit_price=80,
                tax_rate=0,
                line_total=80,
                supply_category="zero",
            ),
            m.SalesInvoiceItem(
                tenant_id=tenant_id,
                sales_invoice_id=inv.id,
                product_id=product_ex.id,
                quantity=1,
                unit_price=50,
                tax_rate=0,
                line_total=50,
                supply_category="exempt",
            ),
        ]
    )
    # POS mix: standard + exempt
    db_session.add(
        m.Transaction(
            tenant_id=tenant_id,
            tx_type="pos_sale",
            reference="POS-SPLIT-1",
            subtotal=70,
            tax=3,
            total=73,
            status="completed",
            payload={
                "items": [
                    {
                        "quantity": 1,
                        "unit_price": 20,
                        "tax_rate": 15,
                        "line_subtotal": 20,
                        "supply_category": "standard",
                    },
                    {
                        "quantity": 1,
                        "unit_price": 50,
                        "tax_rate": 0,
                        "line_subtotal": 50,
                        "supply_category": "exempt",
                    },
                ]
            },
            created_at=datetime(2026, 4, 12),
        )
    )
    await db_session.commit()

    pack = await tax_svc.tax_filing_pack(
        db_session,
        tenant_id,
        from_date=datetime(2026, 4, 1),
        to_date=datetime(2026, 4, 30, 23, 59, 59),
    )
    boxes = pack["filing_boxes"]
    assert boxes["taxable_outputs_net"] == 120.0  # 100 invoice + 20 POS
    assert boxes["zero_rated_outputs_net"] == 80.0
    assert boxes["exempt_outputs_net"] == 100.0  # 50 invoice + 50 POS
    by_code = {b["code"]: b["amount"] for b in boxes["boxes"]}
    assert by_code["taxable_outputs_net"] == 120.0
    assert by_code["zero_rated_outputs_net"] == 80.0
    assert by_code["exempt_outputs_net"] == 100.0


def test_classify_supply_category_rules():
    assert tax_svc.classify_supply_category(tax_exempt=True, rate_pct=15) == "exempt"
    assert tax_svc.classify_supply_category(tax_exempt=False, rate_pct=0) == "zero"
    assert tax_svc.classify_supply_category(tax_exempt=False, rate_pct=15) == "standard"
    assert tax_svc.resolve_line_supply_category(None, tax_rate=0) == "zero"
    assert tax_svc.resolve_line_supply_category("exempt", tax_rate=15) == "exempt"


@pytest.mark.asyncio
async def test_tax_filing_falls_back_to_purchase_orders(db_session, seeded):
    tenant_id = seeded["t1"].id
    supplier = m.Party(tenant_id=tenant_id, name="PO Vendor", kind="supplier", credit_limit=0)
    db_session.add(supplier)
    await db_session.flush()

    po = m.PurchaseOrder(
        tenant_id=tenant_id,
        po_number="PO-TAX-1",
        supplier_id=supplier.id,
        status="received",
        subtotal=50,
        tax_amount=7.5,
        total_amount=57.5,
        created_by=seeded["admin1"].id,
    )
    db_session.add(po)
    await db_session.commit()

    pack = await tax_svc.tax_filing_pack(db_session, tenant_id)
    assert pack["input_tax_source"] == "purchase_orders"
    assert pack["input_tax"] == 7.5
    assert pack["schedules"]["input"][0]["document_number"] == "PO-TAX-1"


def test_exportable_includes_tax_filing():
    assert "tax_filing" in EXPORTABLE


def test_flatten_tax_filing():
    payload = {
        "output_tax": 10,
        "input_tax": 4,
        "net_tax_payable": 6,
        "input_tax_source": "purchase_invoices",
        "filing_boxes": {
            "boxes": [
                {"box": "1", "code": "taxable_outputs_net", "label": "Taxable outputs", "amount": 100},
                {"box": "5", "code": "net_tax_payable", "label": "Net", "amount": 6},
            ]
        },
        "schedules": {
            "output": [{"document_number": "INV-1", "tax_amount": 10}],
            "input": [{"document_number": "PI-1", "tax_amount": 4}],
        },
        "lines": [],
    }
    rows, lines, title = flatten_report("tax_filing", payload)
    assert title == "Tax Filing Pack"
    assert any(r.get("section") == "filing_box" for r in rows)
    assert any("FILING BOXES" in line for line in lines)


@pytest.mark.asyncio
async def test_tax_filing_xlsx_multisheet(db_session, seeded):
    tenant_id = seeded["t1"].id
    party = m.Party(tenant_id=tenant_id, name="XLSX Buyer", kind="customer", credit_limit=0)
    db_session.add(party)
    await db_session.flush()
    inv = m.SalesInvoice(
        tenant_id=tenant_id,
        invoice_number="INV-XLSX-1",
        customer_id=party.id,
        status="posted",
        subtotal=100,
        tax_amount=15,
        total_amount=115,
        paid_amount=0,
        posted_at=datetime.utcnow(),
        created_by=seeded["admin1"].id,
    )
    db_session.add(inv)
    await db_session.commit()

    raw, media, filename = await export_report(db_session, tenant_id, "tax_filing", "xlsx")
    assert media.startswith("application/vnd.openxmlformats")
    assert filename.endswith(".xlsx")
    wb = load_workbook(BytesIO(raw))
    assert set(wb.sheetnames) == {"Summary", "FilingBoxes", "OutputSchedule", "InputSchedule"}
    assert wb["FilingBoxes"]["A1"].value == "box"
