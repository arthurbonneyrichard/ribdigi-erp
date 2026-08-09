"""Ghana GRA VAT government filing template tests."""

from datetime import datetime
from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app import models as m
from app import tax as tax_svc
from app import tax_filings as tax_filings_svc
from app.report_export import EXPORTABLE, export_report, flatten_report


def test_tax_filing_gh_exportable():
    assert "tax_filing_gh" in EXPORTABLE


@pytest.mark.asyncio
async def test_gh_vat_boxes_match_neutral_pack(db_session, seeded):
    tenant_id = seeded["t1"].id
    tenant = await db_session.get(m.Tenant, tenant_id)
    tenant.tax_jurisdiction = "GH"
    tenant.tax_registration_number = "C0001234567"
    party = m.Party(tenant_id=tenant_id, name="Buyer", kind="customer", credit_limit=0)
    supplier = m.Party(tenant_id=tenant_id, name="Vendor", kind="supplier", credit_limit=0)
    db_session.add_all([party, supplier])
    await db_session.flush()

    inv = m.SalesInvoice(
        tenant_id=tenant_id,
        invoice_number="INV-GH-1",
        customer_id=party.id,
        status="posted",
        subtotal=200,
        tax_amount=30,
        total_amount=230,
        paid_amount=0,
        posted_at=datetime(2026, 3, 15),
        created_by=seeded["admin1"].id,
    )
    pi = m.PurchaseInvoice(
        tenant_id=tenant_id,
        invoice_number="PI-GH-1",
        supplier_id=supplier.id,
        status="unpaid",
        subtotal=100,
        tax_amount=15,
        total_amount=115,
        paid_amount=0,
        invoice_date=datetime(2026, 3, 10),
        created_by=seeded["admin1"].id,
    )
    db_session.add_all([inv, pi])
    await db_session.commit()

    pack = await tax_filings_svc.government_filing_pack(
        db_session,
        tenant_id,
        from_date=datetime(2026, 3, 1),
        to_date=datetime(2026, 3, 31, 23, 59, 59),
        jurisdiction="GH",
    )
    gov = pack["government"]
    assert gov["template"] == "gh_vat_return"
    assert gov["header"]["tax_registration_number"] == "C0001234567"
    by_code = {b["code"]: b["amount"] for b in gov["boxes"]}
    assert by_code["GH1"] == pack["filing_boxes"]["taxable_outputs_net"]
    assert by_code["GH3"] == pack["filing_boxes"]["zero_rated_outputs_net"]
    assert by_code["GH4"] == pack["filing_boxes"]["exempt_outputs_net"]
    assert by_code["GH5"] == pack["output_tax"]
    assert by_code["GH8"] == pack["input_tax"]
    assert by_code["GH9"] == pack["net_tax_payable"]
    assert by_code["GH9"] == 15.0


@pytest.mark.asyncio
async def test_gh_vat_maps_zero_and_exempt_boxes(db_session, seeded):
    tenant_id = seeded["t1"].id
    tenant = await db_session.get(m.Tenant, tenant_id)
    tenant.tax_jurisdiction = "GH"
    tenant.tax_registration_number = "C000999"
    party = m.Party(tenant_id=tenant_id, name="GH Split Buyer", kind="customer", credit_limit=0)
    product = m.Product(
        tenant_id=tenant_id,
        name="Mix",
        sku="GH-MIX",
        selling_price=10,
        cost_price=1,
    )
    db_session.add_all([party, product])
    await db_session.flush()
    inv = m.SalesInvoice(
        tenant_id=tenant_id,
        invoice_number="INV-GH-SPLIT",
        customer_id=party.id,
        status="posted",
        subtotal=150,
        tax_amount=15,
        total_amount=165,
        paid_amount=0,
        posted_at=datetime(2026, 5, 5),
        created_by=seeded["admin1"].id,
    )
    db_session.add(inv)
    await db_session.flush()
    db_session.add_all(
        [
            m.SalesInvoiceItem(
                tenant_id=tenant_id,
                sales_invoice_id=inv.id,
                product_id=product.id,
                quantity=1,
                unit_price=100,
                tax_rate=15,
                line_total=115,
                supply_category="standard",
            ),
            m.SalesInvoiceItem(
                tenant_id=tenant_id,
                sales_invoice_id=inv.id,
                product_id=product.id,
                quantity=1,
                unit_price=30,
                tax_rate=0,
                line_total=30,
                supply_category="zero",
            ),
            m.SalesInvoiceItem(
                tenant_id=tenant_id,
                sales_invoice_id=inv.id,
                product_id=product.id,
                quantity=1,
                unit_price=20,
                tax_rate=0,
                line_total=20,
                supply_category="exempt",
            ),
        ]
    )
    await db_session.commit()

    pack = await tax_filings_svc.government_filing_pack(
        db_session,
        tenant_id,
        from_date=datetime(2026, 5, 1),
        to_date=datetime(2026, 5, 31, 23, 59, 59),
        jurisdiction="GH",
    )
    by_code = {b["code"]: b["amount"] for b in pack["government"]["boxes"]}
    assert by_code["GH1"] == 100.0
    assert by_code["GH3"] == 30.0
    assert by_code["GH4"] == 20.0


@pytest.mark.asyncio
async def test_gh_vat_warns_without_tin(db_session, seeded):
    tenant = seeded["t1"]
    tenant.tax_registration_number = None
    await db_session.commit()
    pack = await tax_svc.tax_filing_pack(db_session, tenant.id)
    gov = tax_filings_svc.build_government_return(pack, tenant, jurisdiction="GH")
    assert gov["header"]["tin_missing"] is True
    assert any("TIN" in w for w in gov["warnings"])


@pytest.mark.asyncio
async def test_unsupported_jurisdiction_rejected(db_session, seeded):
    with pytest.raises(HTTPException) as exc:
        await tax_filings_svc.government_filing_pack(
            db_session, seeded["t1"].id, jurisdiction="XX"
        )
    assert exc.value.status_code == 400


@pytest.mark.asyncio
async def test_tax_filing_gh_xlsx_sheets(db_session, seeded):
    tenant_id = seeded["t1"].id
    tenant = await db_session.get(m.Tenant, tenant_id)
    tenant.tax_registration_number = "C999"
    party = m.Party(tenant_id=tenant_id, name="X Buyer", kind="customer", credit_limit=0)
    db_session.add(party)
    await db_session.flush()
    db_session.add(
        m.SalesInvoice(
            tenant_id=tenant_id,
            invoice_number="INV-GH-X",
            customer_id=party.id,
            status="posted",
            subtotal=100,
            tax_amount=15,
            total_amount=115,
            paid_amount=0,
            posted_at=datetime.utcnow(),
            created_by=seeded["admin1"].id,
        )
    )
    await db_session.commit()

    raw, media, filename = await export_report(db_session, tenant_id, "tax_filing_gh", "xlsx")
    assert media.startswith("application/vnd.openxmlformats")
    assert filename.startswith("tax_filing_gh_")
    wb = load_workbook(BytesIO(raw))
    assert set(wb.sheetnames) == {"ReturnHeader", "GHBoxes", "OutputSchedule", "InputSchedule"}
    assert wb["GHBoxes"]["A1"].value == "box"


def test_flatten_tax_filing_gh():
    payload = {
        "government": {
            "template": "gh_vat_return",
            "template_name": "Ghana GRA VAT Return",
            "header": {"taxpayer_name": "Acme", "tax_registration_number": "T1"},
            "boxes": [{"box": "9", "code": "GH9", "label": "Net", "amount": 10}],
            "schedules": {"output": [], "input": []},
            "warnings": [],
        }
    }
    rows, lines, title = flatten_report("tax_filing_gh", payload)
    assert title == "Ghana GRA VAT Return"
    assert any(r.get("section") == "gh_box" for r in rows)
    assert any("GH VAT BOXES" in line for line in lines)
