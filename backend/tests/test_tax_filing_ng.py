"""Nigeria FIRS VAT government filing template tests."""

from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app import models as m
from app import tax as tax_svc
from app import tax_filings as tax_filings_svc
from app.report_export import EXPORTABLE, export_report, flatten_report


def test_tax_filing_ng_exportable():
    assert "tax_filing_ng" in EXPORTABLE
    codes = {j["jurisdiction"] for j in tax_filings_svc.list_supported()}
    assert "NG" in codes
    assert "GH" in codes


@pytest.mark.asyncio
async def test_ng_vat_boxes_match_neutral_pack(db_session, seeded):
    tenant_id = seeded["t1"].id
    tenant = await db_session.get(m.Tenant, tenant_id)
    tenant.tax_jurisdiction = "NG"
    tenant.currency = "NGN"
    tenant.tax_registration_number = "12345678-0001"
    party = m.Party(tenant_id=tenant_id, name="NG Buyer", kind="customer", credit_limit=0)
    supplier = m.Party(tenant_id=tenant_id, name="NG Vendor", kind="supplier", credit_limit=0)
    db_session.add_all([party, supplier])
    await db_session.flush()

    inv = m.SalesInvoice(
        tenant_id=tenant_id,
        invoice_number="INV-NG-1",
        customer_id=party.id,
        status="posted",
        subtotal=200,
        tax_amount=15,
        total_amount=215,
        paid_amount=0,
        posted_at=datetime(2026, 6, 10),
        created_by=seeded["admin1"].id,
    )
    pi = m.PurchaseInvoice(
        tenant_id=tenant_id,
        invoice_number="PI-NG-1",
        supplier_id=supplier.id,
        status="unpaid",
        subtotal=80,
        tax_amount=6,
        total_amount=86,
        paid_amount=0,
        invoice_date=datetime(2026, 6, 5),
        created_by=seeded["admin1"].id,
    )
    db_session.add_all([inv, pi])
    await db_session.commit()

    pack = await tax_filings_svc.government_filing_pack(
        db_session,
        tenant_id,
        from_date=datetime(2026, 6, 1),
        to_date=datetime(2026, 6, 30, 23, 59, 59),
        jurisdiction="NG",
    )
    gov = pack["government"]
    assert gov["template"] == "ng_vat_return"
    assert gov["jurisdiction"] == "NG"
    assert gov["header"]["tax_registration_number"] == "12345678-0001"
    assert gov["header"]["currency"] == "NGN"
    by_code = {b["code"]: b["amount"] for b in gov["boxes"]}
    assert by_code["NG1"] == pack["filing_boxes"]["taxable_outputs_net"]
    assert by_code["NG3"] == pack["filing_boxes"]["zero_rated_outputs_net"]
    assert by_code["NG4"] == pack["filing_boxes"]["exempt_outputs_net"]
    assert by_code["NG5"] == pack["output_tax"]
    assert by_code["NG8"] == pack["input_tax"]
    assert by_code["NG9"] == pack["net_tax_payable"]
    assert by_code["NG9"] == 9.0


@pytest.mark.asyncio
async def test_ng_vat_maps_supply_splits(db_session, seeded):
    tenant_id = seeded["t1"].id
    tenant = await db_session.get(m.Tenant, tenant_id)
    tenant.tax_jurisdiction = "NG"
    tenant.tax_registration_number = "NG-TIN-1"
    party = m.Party(tenant_id=tenant_id, name="NG Split Buyer", kind="customer", credit_limit=0)
    product = m.Product(
        tenant_id=tenant_id,
        name="NG Mix",
        sku="NG-MIX",
        selling_price=10,
        cost_price=1,
    )
    db_session.add_all([party, product])
    await db_session.flush()
    inv = m.SalesInvoice(
        tenant_id=tenant_id,
        invoice_number="INV-NG-SPLIT",
        customer_id=party.id,
        status="posted",
        subtotal=150,
        tax_amount=7.5,
        total_amount=157.5,
        paid_amount=0,
        posted_at=datetime(2026, 7, 8),
        created_by=seeded["admin1"].id,
    )
    db_session.add(inv)
    await db_session.flush()
    db_session.add_all(
        [
            m.SalesInvoiceItem(
                tenant_id=tenant_id,
                sales_invoice_id=inv.id,
                product_id=product.id,
                quantity=1,
                unit_price=100,
                tax_rate=7.5,
                line_total=107.5,
                supply_category="standard",
            ),
            m.SalesInvoiceItem(
                tenant_id=tenant_id,
                sales_invoice_id=inv.id,
                product_id=product.id,
                quantity=1,
                unit_price=40,
                tax_rate=0,
                line_total=40,
                supply_category="zero",
            ),
            m.SalesInvoiceItem(
                tenant_id=tenant_id,
                sales_invoice_id=inv.id,
                product_id=product.id,
                quantity=1,
                unit_price=10,
                tax_rate=0,
                line_total=10,
                supply_category="exempt",
            ),
        ]
    )
    await db_session.commit()

    pack = await tax_filings_svc.government_filing_pack(
        db_session,
        tenant_id,
        from_date=datetime(2026, 7, 1),
        to_date=datetime(2026, 7, 31, 23, 59, 59),
        jurisdiction="NG",
    )
    by_code = {b["code"]: b["amount"] for b in pack["government"]["boxes"]}
    assert by_code["NG1"] == 100.0
    assert by_code["NG3"] == 40.0
    assert by_code["NG4"] == 10.0


@pytest.mark.asyncio
async def test_ng_vat_warns_without_tin(db_session, seeded):
    tenant = seeded["t1"]
    tenant.tax_registration_number = None
    await db_session.commit()
    pack = await tax_svc.tax_filing_pack(db_session, tenant.id)
    gov = tax_filings_svc.build_government_return(pack, tenant, jurisdiction="NG")
    assert gov["header"]["tin_missing"] is True
    assert any("TIN" in w for w in gov["warnings"])


@pytest.mark.asyncio
async def test_tax_filing_ng_xlsx_sheets(db_session, seeded):
    tenant_id = seeded["t1"].id
    tenant = await db_session.get(m.Tenant, tenant_id)
    tenant.tax_registration_number = "NG999"
    party = m.Party(tenant_id=tenant_id, name="NG X Buyer", kind="customer", credit_limit=0)
    db_session.add(party)
    await db_session.flush()
    db_session.add(
        m.SalesInvoice(
            tenant_id=tenant_id,
            invoice_number="INV-NG-X",
            customer_id=party.id,
            status="posted",
            subtotal=100,
            tax_amount=7.5,
            total_amount=107.5,
            paid_amount=0,
            posted_at=datetime.utcnow(),
            created_by=seeded["admin1"].id,
        )
    )
    await db_session.commit()

    raw, media, filename = await export_report(db_session, tenant_id, "tax_filing_ng", "xlsx")
    assert media.startswith("application/vnd.openxmlformats")
    assert filename.startswith("tax_filing_ng_")
    wb = load_workbook(BytesIO(raw))
    assert set(wb.sheetnames) == {"ReturnHeader", "NGBoxes", "OutputSchedule", "InputSchedule"}
    assert wb["NGBoxes"]["A1"].value == "box"


def test_flatten_tax_filing_ng():
    payload = {
        "government": {
            "template": "ng_vat_return",
            "template_name": "Nigeria FIRS VAT Return",
            "header": {"taxpayer_name": "Lagos Co", "tax_registration_number": "T1"},
            "boxes": [{"box": "9", "code": "NG9", "label": "Net", "amount": 10}],
            "schedules": {"output": [], "input": []},
            "warnings": [],
        }
    }
    rows, lines, title = flatten_report("tax_filing_ng", payload)
    assert title == "Nigeria FIRS VAT Return"
    assert any(r.get("section") == "ng_box" for r in rows)
    assert any("NG VAT BOXES" in line for line in lines)
