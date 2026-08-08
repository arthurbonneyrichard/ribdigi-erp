from io import BytesIO

from openpyxl import load_workbook

from app.report_export import to_csv, to_pdf, to_xlsx, flatten_report, EXPORTABLE, EXPORT_FORMATS


def test_to_csv_roundtrip():
    text = to_csv([{"sku": "A1", "qty": 2}, {"sku": "B2", "qty": 3}])
    assert "sku,qty" in text.replace(" ", "")
    assert "A1" in text
    assert "B2" in text


def test_to_csv_empty_with_headers():
    text = to_csv([], fieldnames=["a", "b"])
    assert text.startswith("a,b")


def test_pdf_has_header_and_eof():
    raw = to_pdf("Trial Balance", ["Cash: 100", "Sales: 50"], subtitle="Acme")
    assert raw.startswith(b"%PDF-1.4")
    assert b"%%EOF" in raw
    assert b"Trial Balance" in raw


def test_to_xlsx_roundtrip():
    raw = to_xlsx(
        [{"sku": "A1", "qty": 2}, {"sku": "B2", "qty": 3}],
        sheet_title="Inventory Balance",
    )
    assert raw[:2] == b"PK"  # zip/xlsx magic
    wb = load_workbook(BytesIO(raw))
    ws = wb.active
    assert ws.title == "Inventory Balance"
    assert [c.value for c in ws[1]] == ["sku", "qty"]
    assert ws["A2"].value == "A1"
    assert ws["B3"].value == 3


def test_to_xlsx_empty_headers():
    raw = to_xlsx([], fieldnames=["a", "b"], sheet_title="Empty")
    wb = load_workbook(BytesIO(raw))
    assert [c.value for c in wb.active[1]] == ["a", "b"]


def test_export_formats_include_xlsx():
    assert EXPORT_FORMATS == frozenset({"csv", "pdf", "xlsx"})


def test_flatten_balance_sheet():
    payload = {
        "total_assets": 100,
        "total_liabilities": 40,
        "total_equity": 60,
        "balanced": True,
        "assets": [{"code": "1000", "name": "Cash", "balance": 100}],
        "liabilities": [{"code": "2000", "name": "AP", "balance": 40}],
        "equity": [{"code": "3000", "name": "Equity", "balance": 60}],
    }
    rows, lines, title = flatten_report("balance_sheet", payload)
    assert title == "Balance Sheet"
    assert any(r.get("section") == "assets" for r in rows)
    assert any("ASSETS" in line for line in lines)


def test_exportable_includes_balance_sheet():
    assert "balance_sheet" in EXPORTABLE
    assert "sales_products" in EXPORTABLE


def test_flatten_sales_products():
    rows, lines, title = flatten_report(
        "sales_products",
        {"products": [{"sku": "X", "name": "Widget", "quantity": 1, "revenue": 10}], "total_revenue": 10},
    )
    assert rows[0]["sku"] == "X"
    assert "Sales by Product" == title
