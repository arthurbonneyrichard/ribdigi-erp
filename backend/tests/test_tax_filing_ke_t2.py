"""Stage 10 T2: Kenya KRA VAT government filing template."""

from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app import models as m
from app import tax_filings as tax_filings_svc
from app.report_export import EXPORTABLE, export_report, flatten_report


def test_tax_filing_ke_exportable():
    assert "tax_filing_ke" in EXPORTABLE
    supported = {row["jurisdiction"] for row in tax_filings_svc.list_supported()}
    assert "KE" in supported


@pytest.mark.asyncio
async def test_ke_vat_boxes_match_neutral_pack(db_session, seeded):
    tenant_id = seeded["t1"].id
    tenant = await db_session.get(m.Tenant, tenant_id)
    tenant.tax_jurisdiction = "KE"
    tenant.tax_registration_number = "P051234567A"
    party = m.Party(tenant_id=tenant_id, name="Buyer KE", kind="customer", credit_limit=0)
    supplier = m.Party(tenant_id=tenant_id, name="Vendor KE", kind="supplier", credit_limit=0)
    db_session.add_all([party, supplier])
    await db_session.flush()

    inv = m.SalesInvoice(
        tenant_id=tenant_id,
        invoice_number="INV-KE-1",
        customer_id=party.id,
        status="posted",
        subtotal=200,
        tax_amount=32,
        total_amount=232,
        paid_amount=0,
        posted_at=datetime(2026, 3, 15),
        created_by=seeded["admin1"].id,
    )
    pi = m.PurchaseInvoice(
        tenant_id=tenant_id,
        invoice_number="PI-KE-1",
        supplier_id=supplier.id,
        status="unpaid",
        subtotal=100,
        tax_amount=16,
        total_amount=116,
        paid_amount=0,
        invoice_date=datetime(2026, 3, 10),
        created_by=seeded["admin1"].id,
    )
    db_session.add_all([inv, pi])
    await db_session.commit()

    pack = await tax_filings_svc.government_filing_pack(
        db_session,
        tenant_id,
        from_date=datetime(2026, 3, 1),
        to_date=datetime(2026, 3, 31, 23, 59, 59),
        jurisdiction="KE",
    )
    gov = pack["government"]
    assert gov["template"] == "ke_vat_return"
    assert gov["jurisdiction"] == "KE"
    assert gov["header"]["tax_registration_number"] == "P051234567A"
    by_code = {b["code"]: b["amount"] for b in gov["boxes"]}
    assert by_code["KE1"] == pack["filing_boxes"]["taxable_outputs_net"]
    assert by_code["KE5"] == pack["output_tax"]
    assert by_code["KE8"] == pack["input_tax"]
    assert by_code["KE9"] == pack["net_tax_payable"]
    assert by_code["KE9"] == pytest.approx(16.0)
    assert any("e-file" in w.lower() or "itax" in w.lower() for w in gov["warnings"])


@pytest.mark.asyncio
async def test_tax_filing_ke_xlsx_sheets(db_session, seeded):
    tenant_id = seeded["t1"].id
    tenant = await db_session.get(m.Tenant, tenant_id)
    tenant.tax_registration_number = "P051111111A"
    party = m.Party(tenant_id=tenant_id, name="KE X Buyer", kind="customer", credit_limit=0)
    db_session.add(party)
    await db_session.flush()
    db_session.add(
        m.SalesInvoice(
            tenant_id=tenant_id,
            invoice_number="INV-KE-X",
            customer_id=party.id,
            status="posted",
            subtotal=100,
            tax_amount=16,
            total_amount=116,
            paid_amount=0,
            posted_at=datetime.utcnow(),
            created_by=seeded["admin1"].id,
        )
    )
    await db_session.commit()

    raw, media, filename = await export_report(db_session, tenant_id, "tax_filing_ke", "xlsx")
    assert media.startswith("application/vnd.openxmlformats")
    assert filename.startswith("tax_filing_ke_")
    wb = load_workbook(BytesIO(raw))
    assert set(wb.sheetnames) == {"ReturnHeader", "KEBoxes", "OutputSchedule", "InputSchedule"}
    assert wb["KEBoxes"]["A1"].value == "box"


def test_flatten_tax_filing_ke():
    payload = {
        "government": {
            "template": "ke_vat_return",
            "template_name": "Kenya KRA VAT Return",
            "header": {"taxpayer_name": "Acme KE", "tax_registration_number": "P05"},
            "boxes": [{"box": "9", "code": "KE9", "label": "Net", "amount": 10}],
            "schedules": {"output": [], "input": []},
            "warnings": ["Manual KRA filing workbook only — does not e-file to the KRA iTax portal"],
        }
    }
    rows, lines, title = flatten_report("tax_filing_ke", payload)
    assert title == "Kenya KRA VAT Return"
    assert any(r.get("section") == "ke_box" for r in rows)
    assert any("KE VAT BOXES" in line for line in lines)
    assert any("e-file" in line.lower() or "itax" in line.lower() for line in lines)
